# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

dataStr = 'data'
with open(dataStr + '.json', 'r', encoding='utf-8') as file:
    json_data = json.load(file)

# 创建工作簿
workbook = Workbook()
# 获取活动工作表
worksheet = workbook.active
worksheet.title = "成员信息"

# 定义表头数据
headers = list(json_data[0].keys()) if json_data else []
# 修正扩展解包语法
worksheet.append(["序号", *headers])

# 设置表头样式
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
header_font = Font(bold=True, size=14, color="000000")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_border = Border(
    top=Side(style="thin"),
    left=Side(style="thin"),
    bottom=Side(style="thin"),
    right=Side(style="thin")
)
# 应用表头样式到每一个表头单元格，包含序号列
for col_num in range(1, len(headers) + 2):
    cell = worksheet.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = header_border

# 填充数据，使用 enumerate 获取索引
for index, row_data in enumerate(json_data, start=1):
    # 修正扩展解包语法
    row_values = [index, *list(row_data.values())]
    worksheet.append(row_values)

def adjust_column_width(ws):
    max_lengths = {}
    # 包含序号列
    headers_with_index = ["序号", *headers]
    for i, header in enumerate(headers_with_index):
        col_letter = get_column_letter(i + 1)
        max_lengths[col_letter] = {'value': len(header), 'type': 'str'}
    for row in ws.iter_rows():
        for cell in row:
            col_letter = cell.column_letter
            try:
                if isinstance(cell.value, (int, float)):  # 如果是数字
                    max_lengths[col_letter]['type'] = 'int'
                max_lengths[col_letter]['value'] = max(max_lengths[col_letter]['value'], len(str(cell.value)))
            except:
                max_lengths[col_letter] = {'value': len(str(cell.value)), 'type': 'str'}
    for col, length in max_lengths.items():
        if length['type'] == 'int':
            ws.column_dimensions[col].width = length['value'] + 4  # 增加数字列的宽度
        else:
            ws.column_dimensions[col].width = length['value'] * 2.5 + 2

adjust_column_width(worksheet)

filename = "联赛参与名单.xlsx"
# 保存工作簿
workbook.save(filename)
print("成功保存文件：" + filename)