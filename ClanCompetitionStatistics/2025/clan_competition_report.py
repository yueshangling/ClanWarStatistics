# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
file1='7月'
file2='8月'
with open('./' + file1 + '/联赛参与名单.json', 'r', encoding='utf-8') as f:
    data1 = json.load(f)
with open('./' + file2 + '/联赛参与名单.json', 'r', encoding='utf-8') as f:
    data2 = json.load(f)

# 创建字典方便查找
data1_dict = {item['名称']: {'繁荣度': int(item['繁荣度']), '大本': item['大本']} for item in data1}
data2_dict = {item['名称']: {'繁荣度': int(item['繁荣度']), '大本': item['大本']} for item in data2}

# 找出只在3月出现的成员
only_in_data1 = [name for name in data1_dict if name not in data2_dict]
# 找出只在4月出现的成员
only_in_data2 = [name for name in data2_dict if name not in data1_dict]

print("只在" + file1 + "出现的成员:")
for name in only_in_data1:
    print(f"{name}: 繁荣度 {data1_dict[name]['繁荣度']}, 大本 {data1_dict[name]['大本']}")

print("只在" + file2 + "出现的成员:")
for name in only_in_data2:
    print(f"{name}: 繁荣度 {data2_dict[name]['繁荣度']}, 大本 {data2_dict[name]['大本']}")

# ... existing code ...
# 计算繁荣度变化并包含大本等级
result = []
for name, info2 in data2_dict.items():
    if name in data1_dict:
        change = info2['繁荣度'] - data1_dict[name]['繁荣度']
        result.append({
            '名称': name,
            '原繁荣度': data1_dict[name]['繁荣度'],
            '新繁荣度': info2['繁荣度'],
            '变化量': change,
            '原大本等级': data1_dict[name]['大本'],
            '新大本等级': info2['大本']
        })
# 创建工作簿
workbook = Workbook()
# 处理我方队员进攻数据的工作表
worksheet_player = workbook.active
worksheet_player.title = "繁荣度统计"
# 定义表头数据
headers_player = ["序号", "名称", "原繁荣度", "新繁荣度",'繁荣度变化量','原大本等级','新大本等级' ]
worksheet_player.append(headers_player)

# 设置表头样式
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
header_font = Font(bold=True, size=14, color="000000")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_border = Border(
    top=Side(style="thin"),
    left=Side(style="thin"),
    bottom=Side(style="thin"),
    right=Side(style="thin")
)
# 应用表头样式到每一个表头单元格
for col_num, col_header in enumerate(headers_player, start=1):
    cell = worksheet_player.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = header_border

# 填充我方队员进攻数据，并设置数据单元格样式
for index, data in enumerate(result, start=1):
    row_data = [index, data["名称"], data["原繁荣度"], data.get("新繁荣度", ""),data.get("变化量", ""),data.get("原大本等级", ""),data.get("新大本等级", "")]
    worksheet_player.append(row_data)
def adjust_column_width(ws):
    max_lengths = {}
    for i, header in enumerate(headers_player):
        col_letter = chr(65 + i)
        max_lengths[col_letter] = {'value': len(header), 'type': 'str'}
    for row in ws.iter_rows():
        for cell in row:
            col_letter = cell.column_letter
            try:
                if isinstance(cell.value, float):  # 如果是数字
                    max_lengths[col_letter]['type'] = 'int'
                max_lengths[col_letter]['value'] = max(max_lengths[col_letter]['value'], len(str(cell.value)))
            except:
                max_lengths[col_letter] = {'value': len(str(cell.value)), 'type': 'str'}
    for col, length in max_lengths.items():
        if length['type'] == 'int':
            ws.column_dimensions[col].width = length['value'] + 4  # 增加数字列的宽度
        else:
            ws.column_dimensions[col].width = length['value'] * 2.5 + 2
adjust_column_width(worksheet_player)
date_str = file1 + '-' + file2 + '繁荣度变化'
filename =  date_str + ".xlsx"
# 保存工作簿
workbook.save(filename)
print("成功保存文件：" + filename)
