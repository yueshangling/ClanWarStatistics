# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import json
with open('./data.json', 'r', encoding='utf-8') as file:
    game_data = json.load(file)
# 创建工作簿
workbook = Workbook()
# 处理我方队员进攻数据的工作表
worksheet_player = workbook.active
worksheet_player.title = "我方队员进攻"
# 定义表头数据
type = 2
if type == 1:
    headers_player = ["序号", "名称", "职位", "部落等级", "第一次攻击", "第一次攻击详情", "第二次攻击", "第二次攻击详情", ""]
else:
    headers_player = ["序号", "名称", "职位", "部落等级", "第一次攻击", "第一次攻击详情", "获得的星", "评价"]
worksheet_player.append(headers_player)

# 设置表头样式
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
header_font = Font(bold=True, size=14, color="000000")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_border = Border(
    top=Side(style="thin"),
    left=Side(style="thin"),
    bottom=Side(style="thin"),
    right=Side(style="thin")
)
# 应用表头样式到每一个表头单元格
for col_num, col_header in enumerate(headers_player, start=1):
    cell = worksheet_player.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = header_border

# 填充我方队员进攻数据，并设置数据单元格样式
for index, data in enumerate(game_data, start=1):
    if type == 1:
        row_data = [index, data["名称"], data["职位"], data.get("部落等级", "已退出"), data["第一次攻击"],
            data["第一次攻击详情"], data["第二次攻击"], data["第二次攻击详情"], data.get("评价", "")]
    else:
        row_data = [index, data["名称"], data["职位"], data.get("部落等级", "已退出"), data["第一次攻击"],
            data["第一次攻击详情"], data.get("获得的星", ""), data.get("评价", "")]
    worksheet_player.append(row_data)
    for col_num in range(1, len(row_data) + 1):
        cell = worksheet_player.cell(row=worksheet_player.max_row, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # 判断是否两个字段值都是"未使用"，若是则整行标红
        if (data["第一次攻击详情"] == "未使用" and type == 2) or (type == 1 and data["第二次攻击详情"] == "未使用" and data["第一次攻击详情"] == "未使用"):
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for col in range(1, len(headers_player) + 1):
                target_cell = worksheet_player.cell(row=worksheet_player.max_row, column=col)
                target_cell.fill = red_fill
        cell.border = Border(
            top=Side(style="thin"),
            left=Side(style="thin"),
            bottom=Side(style="thin"),
            right=Side(style="thin")
        )
def adjust_column_width(ws):
    max_lengths = {}
    for i, header in enumerate(headers_player):
        col_letter = chr(65 + i)
        max_lengths[col_letter] = {'value': len(header), 'type': 'str'}
    for row in ws.iter_rows():
        for cell in row:
            col_letter = cell.column_letter
            try:
                if isinstance(cell.value, float):  # 如果是数字
                    max_lengths[col_letter]['type'] = 'int'
                max_lengths[col_letter]['value'] = max(max_lengths[col_letter]['value'], len(str(cell.value)))
            except:
                max_lengths[col_letter] = {'value': len(str(cell.value)), 'type': 'str'}
    for col, length in max_lengths.items():
        if length['type'] == 'int':
            ws.column_dimensions[col].width = length['value'] + 4  # 增加数字列的宽度
        else:
            ws.column_dimensions[col].width = length['value'] * 2.5 + 2
adjust_column_width(worksheet_player)
now = datetime.datetime.now()
date_str = now.strftime("%Y-%m-%d")
filename =  date_str + ".xlsx"
# 保存工作簿
workbook.save(filename)
print("成功保存文件：" + filename)
