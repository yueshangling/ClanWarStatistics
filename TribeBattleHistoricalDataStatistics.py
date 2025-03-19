import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from collections import defaultdict

# 用户选择攻击模式
mode = int(input("请选择攻击模式（1为单次攻击，2为两次攻击）："))

headers_player = [
    "名称",
    "1星",
    "1星占比",
    "2星",
    "2星占比",
    "3星",
    "3星占比",
    "黑三",
    "黑三占比",
    "总进攻次数",
    "总获得星星",  # 添加缺少的逗号
    "未使用进攻次数",
    "未使用进攻次数占比",
    "第一次攻击占比",
]
if mode == 2:
    headers_player.append("第二次攻击占比")

def collect_unused_attacks(folder_path, mode):
    """收集未使用的攻击记录"""
    dir_unused = defaultdict(list)
    total_unused = []
    
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.json'):
                file_path = os.path.join(root, file)
                with open(file_path, 'r', encoding='utf-8') as f:
                    try:
                        data = json.load(f)
                        for entry in data:
                            relative_path = os.path.relpath(root, folder_path)
                            # 根据模式判断条件
                            if mode == 2:
                                first = entry.get("第一次攻击详情", "未使用")
                                second = entry.get("第二次攻击详情", "未使用")
                                if first == "未使用" and second == "未使用":
                                    record = {
                                        "名称": entry.get("名称", "未知"),
                                        "第一次攻击详情": first,
                                        "第二次攻击详情": second,
                                        "来源文件": os.path.basename(file_path),
                                        "来源目录": relative_path
                                    }
                                    dir_unused[relative_path].append(record)
                                    total_unused.append(record)
                            else:
                                first = entry.get("第一次攻击详情", "未使用")
                                if first == "未使用":
                                    record = {
                                        "名称": entry.get("名称", "未知"),
                                        "第一次攻击详情": first,
                                        "来源文件": os.path.basename(file_path),
                                        "来源目录": relative_path
                                    }
                                    dir_unused[relative_path].append(record)
                                    total_unused.append(record)
                    except json.JSONDecodeError:
                        print(f"文件解析失败: {file_path}")
    return dir_unused, total_unused

def export_unused_records(dir_unused, total_unused, base_folder, mode):
    """导出未使用记录"""
    # 导出目录记录
    for relative_path, records in dir_unused.items():
        if records:
            dir_path = os.path.join(base_folder, relative_path)
            os.makedirs(dir_path, exist_ok=True)
            output_path = os.path.join(dir_path, "未使用记录.xlsx")
            if mode == 2:
                df = pd.DataFrame(records)[["名称", "第一次攻击详情", "第二次攻击详情", "来源文件", "来源目录"]]
            else:
                df = pd.DataFrame(records)[["名称", "第一次攻击详情", "来源文件", "来源目录"]]
            df.to_excel(output_path, index=False)
    
    # 导出总记录
    if total_unused:
        output_path = os.path.join(base_folder, "总未使用记录.xlsx")
        if mode == 2:
            df = pd.DataFrame(total_unused)[["名称", "第一次攻击详情", "第二次攻击详情", "来源文件", "来源目录"]]
        else:
            df = pd.DataFrame(total_unused)[["名称", "第一次攻击详情", "来源文件", "来源目录"]]
        df.to_excel(output_path, index=False)
    else:
        print("没有未使用的进攻记录")

def count_stars(details):
    """统计星星数"""
    if details == "未使用":
        return 0
    return details.count('★')

def analyze_folder_by_name(folder_path, mode):
    stats_by_name = defaultdict(lambda: {
        "total_attacks": 0,
        "unused_attacks": 0,
        "star_counts": defaultdict(int),
        "first_attack_total": 0,
        "second_attack_total": 0,
    })

    def process_file(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            try:
                data = json.load(file)
                for entry in data:
                    name = entry.get("名称", "未知")
                    stats = stats_by_name[name]

                    # 处理第一次攻击
                    stats["first_attack_total"] += 1
                    stats["total_attacks"] += 1
                    first_attack_detail = entry.get("第一次攻击详情", "未使用")
                    if first_attack_detail != "未使用":
                        stars = count_stars(first_attack_detail)
                        stats["star_counts"][stars] += 1
                    else:
                        stats["unused_attacks"] += 1

                    # 处理第二次攻击（仅模式2）
                    if mode == 2:
                        stats["second_attack_total"] += 1
                        stats["total_attacks"] += 1
                        second_attack_detail = entry.get("第二次攻击详情", "未使用")
                        if second_attack_detail != "未使用":
                            stars = count_stars(second_attack_detail)
                            stats["star_counts"][stars] += 1
                        else:
                            stats["unused_attacks"] += 1
            except json.JSONDecodeError:
                print(f"文件解析失败: {file_path}")

    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.json'):
                process_file(os.path.join(root, file))

    # 计算结果
    results = []
    for name, stats in stats_by_name.items():
        total = stats["total_attacks"]
        unused = stats["unused_attacks"]
        first_total = stats["first_attack_total"]
        second_total = stats["second_attack_total"] if mode == 2 else 0

        stars_1 = stats["star_counts"][1]
        stars_2 = stats["star_counts"][2]
        stars_3 = stats["star_counts"][3]
        stars_0 = stats["star_counts"][0]

        result = {
            "名称": name,
            "1星": stars_1,
            "1星占比": f"{stars_1 / total:.2%}" if total else "0.00%",
            "2星": stars_2,
            "2星占比": f"{stars_2 / total:.2%}" if total else "0.00%",
            "3星": stars_3,
            "3星占比": f"{stars_3 / total:.2%}" if total else "0.00%",
            "黑三": stars_0,
            "黑三占比": f"{stars_0 / total:.2%}" if total else "0.00%",
            "总进攻次数": total,
            "总获得星星": stars_1 * 1 + stars_2 * 2 + stars_3 * 3,  # 添加统计值
            "未使用进攻次数": unused,
            "未使用进攻次数占比": f"{unused / total:.2%}" if total else "0.00%",
            "第一次攻击占比": f"{first_total / total:.2%}" if total else "0.00%",
        }
        if mode == 2:
            result["第二次攻击占比"] = f"{second_total / total:.2%}" if total else "0.00%"
        results.append(result)
    return results
def mark_rows_red(ws):
    # 定义红色填充样式
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # 遍历从第二行开始的每一行，从第 10 列到第 11 列
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=11):
        total_attacks_cell, unused_attacks_cell = row
        if total_attacks_cell.value == unused_attacks_cell.value:
            # 获取整行的单元格范围
            full_row = ws[row[0].row]
            for cell in full_row:
                cell.fill = red_fill
def export_to_excel_with_styles(results, output_path, mode):
    """将结果导出到 Excel 并设置样式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "统计结果"

    # 动态表头
    headers = [
        "名称", "1星", "1星占比", "2星", "2星占比", "3星", "3星占比",
        "黑三", "黑三占比", "总进攻次数", "总获得星星","未使用进攻次数", "未使用进攻次数占比",
        "第一次攻击占比"
    ]
    if mode == 2:
        headers.append("第二次攻击占比")
    ws.append(headers)

    # 设置表头样式
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    header_font = Font(bold=True, size=14, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        bottom=Side(style="thin"),
        right=Side(style="thin")
    )
    for col_num, col_header in enumerate(headers_player, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    # 填充数据
    for res in results:
        row = []
        for h in headers:
            if h == "第二次攻击占比" and h not in res:
                row.append(0.0)
            else:
                val = res.get(h, 0)
                if isinstance(val, str) and "%" in val:
                    row.append(float(val.strip("%")) / 100)
                else:
                    row.append(val)
        ws.append(row)

    # 设置百分比格式
    for col in range(1, len(headers)+1):
        if "占比" in ws.cell(row=1, column=col).value:
            for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
                cell[0].number_format = "0.00%"
    mark_rows_red(ws)
    adjust_column_width(ws,headers)
    wb.save(output_path)
    print(f"统计结果已保存至: {output_path}")
def adjust_column_width(ws,headers):
    max_lengths = {}
    for i, header in enumerate(headers):
        col_letter = chr(65 + i)
        max_lengths[col_letter] = {'value': len(header), 'type': 'str'}
    for row in ws.iter_rows():
        for cell in row:
            col_letter = cell.column_letter
            try:
                if isinstance(cell.value, float):  # 如果是数字
                    max_lengths[col_letter]['type'] = 'int'
                max_lengths[col_letter]['value'] = max(max_lengths[col_letter]['value'], len(str(cell.value)))
            except:
                max_lengths[col_letter] = {'value': len(str(cell.value)), 'type': 'str'}
    for col, length in max_lengths.items():
        if length['type'] == 'int':
            ws.column_dimensions[col].width = length['value'] + 4  # 增加数字列的宽度
        else:
            ws.column_dimensions[col].width = length['value'] * 2.5 + 2
# 运行主程序
folder_path = "ClanCompetitionStatistics/2025/2月/"
output_path = "ClanCompetitionStatistics/2025/2月/2025年2月联赛统计结果.xlsx"

results = analyze_folder_by_name(folder_path, mode)
export_to_excel_with_styles(results, output_path, mode)

dir_unused, total_unused = collect_unused_attacks(folder_path, mode)
export_unused_records(dir_unused, total_unused, folder_path, mode)